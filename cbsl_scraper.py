"""
========================================================
  Government Securities Secondary Market Trade Summary
  Scraper — CBSL (pre-Dec 2025) + PDMO/Treasury (Dec 2025+)
========================================================

SETUP — run once in your terminal:
  pip install requests beautifulsoup4 pdfplumber openpyxl pandas lxml

USAGE:
  python cbsl_scraper.py
========================================================
"""

import requests
from bs4 import BeautifulSoup
import pdfplumber
import pandas as pd
import re
import io
import os
from datetime import datetime

# ──────────────────────────────────────────────────────
#  CONFIGURATION
# ──────────────────────────────────────────────────────
PDMO_YEARS   = [2025, 2026]
PDMO_BASE    = "https://www.treasury.gov.lk"
PDMO_URL_TPL = PDMO_BASE + "/web/reports-secondary-market-trade-summary/section/{year}"

CBSL_URL     = "https://www.cbsl.gov.lk/en/press/secondary-market-trade-summary"
CBSL_BASE    = "https://www.cbsl.gov.lk"

MAX_REPORTS  = 60
OUTPUT_DIR   = "cbsl_output"
CSV_FILE     = os.path.join(OUTPUT_DIR, "trade_summary.csv")
EXCEL_FILE   = os.path.join(OUTPUT_DIR, "trade_summary.xlsx")

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    )
}

os.makedirs(OUTPUT_DIR, exist_ok=True)


# ──────────────────────────────────────────────────────
#  STEP 1A — Collect links from PDMO / Treasury
# ──────────────────────────────────────────────────────
def get_pdmo_links():
    links = []
    for year in PDMO_YEARS:
        url = PDMO_URL_TPL.format(year=year)
        print(f"  PDMO {year}: {url}")
        try:
            resp = requests.get(url, headers=HEADERS, timeout=30)
            resp.raise_for_status()
            soup = BeautifulSoup(resp.text, "lxml")
            for a in soup.find_all("a", href=True):
                href = a["href"]
                if "/api/file/" in href:
                    full_url = href if href.startswith("http") else PDMO_BASE + href
                    date_str = extract_date_from_text(a.get_text(strip=True))
                    links.append((date_str or "Unknown", full_url))
        except Exception as e:
            print(f"  WARNING: Could not load PDMO {year}: {e}")
    print(f"  Found {len(links)} PDMO links")
    return links


# ──────────────────────────────────────────────────────
#  STEP 1B — Collect links from old CBSL page
# ──────────────────────────────────────────────────────
def get_cbsl_links():
    links = []
    print(f"  CBSL: {CBSL_URL}")
    try:
        resp = requests.get(CBSL_URL, headers=HEADERS, timeout=30)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "lxml")
        for a in soup.find_all("a", href=True):
            href = a["href"]
            text = a.get_text(strip=True).lower()
            if href.lower().endswith(".pdf") and "pdmo_commencement" not in href.lower():
                if "secondary" in href.lower() or "trade" in text:
                    full_url = href if href.startswith("http") else CBSL_BASE + href
                    date_str = extract_date_from_text(a.get_text(strip=True)) \
                               or extract_date_from_url(href)
                    links.append((date_str or "Unknown", full_url))
    except Exception as e:
        print(f"  WARNING: Could not load CBSL page: {e}")
    print(f"  Found {len(links)} CBSL links")
    return links


# ──────────────────────────────────────────────────────
#  DATE HELPERS
# ──────────────────────────────────────────────────────
def extract_date_from_text(text):
    if not text:
        return None
    checks = [
        (r"\d{1,2}[.\-/]\d{1,2}[.\-/]\d{4}", ["%d.%m.%Y", "%d-%m-%Y", "%d/%m/%Y"]),
        (r"\d{4}[.\-/]\d{1,2}[.\-/]\d{1,2}", ["%Y-%m-%d", "%Y.%m.%d"]),
        (r"\d{1,2}\s+\w+\s+\d{4}",           ["%d %B %Y", "%d %b %Y"]),
    ]
    for pat, fmts in checks:
        m = re.search(pat, text)
        if m:
            raw = m.group(0).strip()
            for fmt in fmts:
                try:
                    return datetime.strptime(raw, fmt).strftime("%Y-%m-%d")
                except ValueError:
                    pass
    return None


def extract_date_from_url(url):
    m = re.search(r"(\d{4})(\d{2})(\d{2})", url)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    return None


# ──────────────────────────────────────────────────────
#  STEP 2 — Download PDF bytes
# ──────────────────────────────────────────────────────
def download_pdf_bytes(url):
    resp = requests.get(url, headers=HEADERS, timeout=60)
    resp.raise_for_status()
    return resp.content


# ──────────────────────────────────────────────────────
#  STEP 3 — Parse PDF → list of dicts
# ──────────────────────────────────────────────────────
_ISIN_RE = re.compile(r"\b(LK[A-Z0-9]{10,12})\b")

def parse_pdf(pdf_bytes, report_date):
    rows = []
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            # Try table extraction
            for table in (page.extract_tables() or []):
                rows.extend(parse_table(table, report_date))
            # Fallback: raw text line parsing
            if not rows:
                text = page.extract_text() or ""
                rows.extend(parse_text_lines(text, report_date))
    return rows


def parse_table(table, report_date):
    rows = []
    if not table:
        return rows
    header_idx = find_header_row(table)
    if header_idx is None:
        return rows
    headers = clean_headers(table[header_idx])
    for row in table[header_idx + 1:]:
        if not row or all(c in (None, "", " ") for c in row):
            continue
        rec = build_record(headers, row, report_date)
        if rec:
            rows.append(rec)
    return rows


def find_header_row(table):
    keywords = {"isin", "yield", "volume", "tenure", "security", "weighted"}
    for idx, row in enumerate(table):
        if not row:
            continue
        text = " ".join(str(c) for c in row if c).lower()
        if sum(1 for kw in keywords if kw in text) >= 2:
            return idx
    return None


def parse_text_lines(text, report_date):
    """
    Parse PDMO-style lines, e.g.:
      1 LKA18226G030 0.50 Tbill 8.03 8.50 8.50 7.98 8.32 16,547 15
    """
    rows = []
    for line in text.splitlines():
        line = line.strip()
        m = _ISIN_RE.search(line)
        if not m:
            continue
        isin = m.group(1)

        sec_type = ("TBond" if re.search(r"\bTBond\b", line, re.I) else
                    "Tbill" if re.search(r"\bTbill\b", line, re.I) else "Unknown")

        floats = []
        for tok in line.split():
            tok_clean = tok.replace(",", "")
            try:
                floats.append(float(tok_clean))
            except ValueError:
                pass

        # Need at least 8 numbers: row_no?, tenure, 5 yields, volume, trades
        if len(floats) < 7:
            continue

        idx = 0
        # If first number is a small integer row counter, skip it
        if floats[0] == int(floats[0]) and floats[0] < 500:
            idx = 1

        try:
            rows.append({
                "Date"              : report_date,
                "ISIN"              : isin,
                "Security_Type"     : sec_type,
                "Tenure"            : _get(floats, idx),
                "Opening_Yield"     : _get(floats, idx + 1),
                "Closing_Yield"     : _get(floats, idx + 2),
                "Highest_Yield"     : _get(floats, idx + 3),
                "Lowest_Yield"      : _get(floats, idx + 4),
                "Weighted_Avg_Yield": _get(floats, idx + 5),
                "Volume"            : _get(floats, idx + 6),
                "Num_Trades"        : _get(floats, idx + 7),
            })
        except Exception:
            pass
    return rows


def _get(lst, i):
    return lst[i] if i < len(lst) else None


_HDR_MAP = {
    "isin"                  : "ISIN",
    "tenure"                : "Tenure",
    "security type"         : "Security_Type",
    "type"                  : "Security_Type",
    "opening yield"         : "Opening_Yield",
    "closing yield"         : "Closing_Yield",
    "highest yield"         : "Highest_Yield",
    "lowest yield"          : "Lowest_Yield",
    "weighted average yield": "Weighted_Avg_Yield",
    "weighted avg yield"    : "Weighted_Avg_Yield",
    "weighted average rate" : "Weighted_Avg_Yield",
    "volume"                : "Volume",
    "no. of trades"         : "Num_Trades",
    "no of trades"          : "Num_Trades",
    "number of trades"      : "Num_Trades",
    "no"                    : "Row_No",
}

def clean_headers(raw):
    out = []
    for h in raw:
        key = re.sub(r"\s+", " ", str(h or "").lower().replace("\n", " "))
        key = key.replace("(%)", "").replace("(rs. mn)", "").replace("(%)", "").strip()
        out.append(_HDR_MAP.get(key, key.title().replace(" ", "_") or f"Col_{len(out)}"))
    return out


def build_record(headers, row, report_date):
    row = list(row) + [None] * max(0, len(headers) - len(row))
    record = {"Date": report_date}
    for col, val in zip(headers, row):
        v = str(val).strip().replace("\n", " ") if val is not None else None
        record[col] = v if v not in ("", "-", "N/A", "n/a", "None") else None

    isin = str(record.get("ISIN", "") or "")
    if not isin.startswith("LK"):
        return None

    for num_col in ["Tenure", "Opening_Yield", "Closing_Yield",
                    "Highest_Yield", "Lowest_Yield",
                    "Weighted_Avg_Yield", "Volume", "Num_Trades"]:
        if record.get(num_col):
            try:
                record[num_col] = float(str(record[num_col]).replace(",", ""))
            except ValueError:
                record[num_col] = None
    return record


# ──────────────────────────────────────────────────────
#  STEP 4 — Save CSV + Excel
# ──────────────────────────────────────────────────────
def save_outputs(all_rows):
    if not all_rows:
        print("\nNo data rows extracted.")
        return

    df = pd.DataFrame(all_rows)
    preferred = ["Date", "ISIN", "Security_Type", "Tenure",
                 "Opening_Yield", "Closing_Yield",
                 "Highest_Yield", "Lowest_Yield",
                 "Weighted_Avg_Yield", "Volume", "Num_Trades"]
    cols = [c for c in preferred if c in df.columns] + \
           [c for c in df.columns if c not in preferred]
    df = df[cols].sort_values(["Date", "Tenure"], ignore_index=True)

    df.to_csv(CSV_FILE, index=False)
    print(f"\nCSV   saved: {CSV_FILE}")

    with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Trade_Summary", index=False)
        ws = writer.sheets["Trade_Summary"]
        from openpyxl.styles import Font, PatternFill, Alignment
        fill = PatternFill("solid", fgColor="1F4E79")
        font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.font      = font
            cell.fill      = fill
            cell.alignment = Alignment(horizontal="center")
        for col_cells in ws.columns:
            w = max((len(str(c.value or "")) for c in col_cells), default=10)
            ws.column_dimensions[col_cells[0].column_letter].width = min(w + 4, 40)

    print(f"Excel saved: {EXCEL_FILE}")
    print(f"\nTotal rows  : {len(df)}")
    print(f"Date range  : {df['Date'].min()}  to  {df['Date'].max()}")
    print(f"Unique ISINs: {df['ISIN'].nunique()}")
    print("\nSample (first 5 rows):")
    print(df.head().to_string(index=False))


# ──────────────────────────────────────────────────────
#  MAIN
# ──────────────────────────────────────────────────────
def main():
    print("=" * 58)
    print("  Government Securities Trade Summary Scraper")
    print("  PDMO (treasury.gov.lk) + CBSL (cbsl.gov.lk)")
    print("=" * 58)

    print("\nCollecting PDF links...")
    all_links = get_pdmo_links() + get_cbsl_links()

    seen, unique = set(), []
    for item in all_links:
        if item[1] not in seen:
            seen.add(item[1])
            unique.append(item)

    unique.sort(key=lambda x: x[0], reverse=True)
    unique = unique[:MAX_REPORTS]
    print(f"\n{len(unique)} PDFs to download (limit = {MAX_REPORTS})")

    all_rows = []
    for i, (date_str, url) in enumerate(unique, 1):
        print(f"\n[{i:02d}/{len(unique)}] {date_str}")
        print(f"  {url}")
        try:
            pdf_bytes = download_pdf_bytes(url)
            rows      = parse_pdf(pdf_bytes, date_str)
            all_rows.extend(rows)
            print(f"  OK — {len(rows)} rows")
        except Exception as e:
            print(f"  FAILED: {e}")

    save_outputs(all_rows)
    print(f"\nDone! Output folder: {os.path.abspath(OUTPUT_DIR)}")


if __name__ == "__main__":
    main()